import json
import io
from datetime import timedelta, datetime
from decimal import Decimal

from django.contrib import admin
from django import forms
from django.utils.html import format_html, escape
from django.urls import reverse, path
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.db.models import Prefetch, Sum, F, ExpressionWrapper, DecimalField, Q
from django.utils import timezone
from django.db import transaction
from django.utils.safestring import mark_safe
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.contrib.auth.admin import UserAdmin as DjangoUserAdmin
from django.core.exceptions import ValidationError as AdminValidationError
from django.db.models.functions import TruncDay, TruncMonth, Coalesce
from django.template import Template, RequestContext

from .models import (
    Category, Product, ShopConfiguration, BannerImage,
    DocumentPost, Order, OrderItem, Customer, SalesReport
)


class BannerImageInline(admin.TabularInline):
    model = BannerImage
    extra = 1


@admin.register(ShopConfiguration)
class ShopConfigurationAdmin(admin.ModelAdmin):
    list_display = ['title', 'phone', 'email']
    inlines = [BannerImageInline]

    def has_add_permission(self, request):
        return not ShopConfiguration.objects.exists()


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ['name', 'slug']
    prepopulated_fields = {'slug': ('name',)}

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('parent')


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = [
        'name', 'code', 'price', 'sale_price_display', 'import_price_display',
        'stock_display', 'new_import_price', 'new_stock', 'action_button',
        'defective_quantity', 'subtract_defective_button',
        'tax_rate', 'after_tax_profit_display'
    ]
    list_editable = ['code', 'price', 'new_import_price', 'new_stock', 'tax_rate', 'defective_quantity']
    readonly_fields = ['import_price', 'stock', 'sale_price']
    list_filter = ['category']
    search_fields = ['name', 'slug']
    prepopulated_fields = {'slug': ('name',)}
    list_per_page = 50

    @admin.display(description=mark_safe("Giá Nhập<br>(VNĐ)"))
    def import_price_display(self, obj):
        return f"{obj.import_price:,.0f}".replace(",", ".")

    @admin.display(description=mark_safe("Giá Bán Cũ<br>(VNĐ)"))
    def sale_price_display(self, obj):
        return f"{obj.sale_price:,.0f}".replace(",", ".")

    @admin.display(description=mark_safe("Số Lượng<br>Tồn Kho"))
    def stock_display(self, obj):
        return f"{obj.stock:,.0f}".replace(",", ".")

    class Media:
        css = {
            'all': ('admin/css/product_admin.css',)
        }

    def after_tax_profit_display(self, obj):
        if obj.price and obj.import_price and obj.price > 0:
            tax_amount = obj.price * (obj.tax_rate / 100)
            profit = obj.price - obj.import_price - tax_amount
            profit_int = int(profit)

            percentage = (profit / obj.price) * 100
            percentage_str = f"{percentage:.1f}%"
            color = "#51cf66" if profit_int >= 0 else "#ff6b6b"
            profit_str = f"{profit_int:,}đ".replace(",", ".")

            return format_html(
                '<span style="color: {}; font-weight: bold;">{} ({})</span>',
                color, profit_str, percentage_str
            )
        return "-"

    after_tax_profit_display.short_description = mark_safe("Lợi nhuận<br>gộp<br>sau thuế")

    def subtract_defective_button(self, obj):
        url = reverse('admin:product_subtract_defective', args=[obj.pk])
        return format_html('<a class="button" href="{}">Trừ</a>', url)

    subtract_defective_button.short_description = mark_safe("TRỪ<br>HÀNG LỖI")

    def action_button(self, obj):
        url = reverse('admin:product_update_data', args=[obj.pk])
        return format_html('<a class="button" href="{}">CHUYỂN</a>', url)

    action_button.short_description = "Cập nhật"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<path:object_id>/update-data/', self.admin_site.admin_view(self.update_data_view),
                 name='product_update_data'),
            path('<path:object_id>/subtract-defective/', self.admin_site.admin_view(self.subtract_defective_view),
                 name='product_subtract_defective'),
        ]
        return custom_urls + urls

    def update_data_view(self, request, object_id):
        product = get_object_or_404(Product, pk=object_id)

        if product.stock != 0:
            messages.error(request,
                           f"Không thể cập nhật: Tồn kho của '{product.name}' đang là {product.stock}, phải bằng 0 mới được phép cập nhật!")
            return redirect('admin:shop_product_changelist')

        if product.new_stock is None or product.new_stock == 0:
            messages.error(request, "Không thể cập nhật: Số lượng mới phải khác 0!")
            return redirect('admin:shop_product_changelist')

        if product.new_import_price is None or product.new_import_price == 0:
            messages.error(request, "Không thể cập nhật: Giá nhập mới phải khác 0!")
            return redirect('admin:shop_product_changelist')

        old_import_price = product.import_price
        new_import_price = product.new_import_price

        try:
            with transaction.atomic():
                backfilled = 0
                if old_import_price is not None:
                    backfilled = OrderItem.objects.filter(
                        product=product,
                        import_price__isnull=True,
                        order__created_at__lte=timezone.now()
                    ).update(import_price=old_import_price)

                product.sale_price = product.price
                if new_import_price and new_import_price > 0:
                    product.import_price = new_import_price
                    product.new_import_price = 0

                product.stock = product.new_stock if product.new_stock is not None else 0
                product.new_stock = 0
                product.save()

                msg = f"Đã cập nhật thành công sản phẩm: {product.name}."
                if backfilled:
                    msg += f" Đã lưu giá nhập cũ vào {backfilled} mục đơn hàng để bảo toàn báo cáo giá vốn."
                messages.success(request, msg)
        except Exception as e:
            messages.error(request, f"Lỗi khi cập nhật sản phẩm: {str(e)}")

        return redirect('admin:shop_product_changelist')

    def subtract_defective_view(self, request, object_id):
        product = get_object_or_404(Product, pk=object_id)
        defective_qty = product.defective_quantity or 0

        if defective_qty == 0:
            messages.error(request, f"Không thể trừ hàng lỗi: Số lượng hàng lỗi của '{product.name}' phải lớn hơn 0!")
            return redirect('admin:shop_product_changelist')

        if product.stock < defective_qty:
            messages.error(request,
                           f"Không thể trừ hàng lỗi: Tồn kho của '{product.name}' là {product.stock}, nhỏ hơn số lượng hàng lỗi {defective_qty}!")
            return redirect('admin:shop_product_changelist')

        try:
            with transaction.atomic():
                product.stock -= defective_qty
                product.defective_quantity = 0
                product.save()
                messages.success(request,
                                 f"Đã trừ thành công {defective_qty} sản phẩm lỗi của '{product.name}'. Tồn kho hiện tại: {product.stock}")
        except Exception as e:
            messages.error(request, f"Lỗi khi trừ hàng lỗi: {str(e)}")

        return redirect('admin:shop_product_changelist')

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)

        def clean_price(self_form):
            price = self_form.cleaned_data.get("price")
            if price is not None and price < 1000:
                raise forms.ValidationError("Giá bán sản phẩm phải lớn hơn 1000 VNĐ!")
            return price

        form.clean_price = clean_price
        return form

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('category')


@admin.register(DocumentPost)
class DocumentPostAdmin(admin.ModelAdmin):
    list_display = ['title', 'created_at']
    list_filter = ['created_at']
    search_fields = ['title', 'slug']
    prepopulated_fields = {'slug': ('title',)}
    readonly_fields = ['created_at']


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    fields = ('product', 'quantity', 'price_display', 'returned_quantity', 'return_item_action')
    readonly_fields = ['product', 'quantity', 'price_display', 'return_item_action']
    can_delete = False

    @admin.display(description="Giá")
    def price_display(self, obj):
        if not obj.pk:
            return "-"
        base_price = obj.price or 0
        discount = obj.discount_per_unit or 0
        final_price = base_price - discount

        base_str = f"{base_price:,.0f}".replace(",", ".") + "đ"
        final_str = f"{final_price:,.0f}".replace(",", ".") + "đ"

        if discount > 0:
            return format_html(
                '<span style="text-decoration: line-through; color: #6c757d;">{}</span> <span style="margin: 0 4px; color: #6c757d;">→</span> <span style="color: #dc3545; font-weight: bold;">{}</span>',
                base_str, final_str
            )
        return base_str

    @admin.display(description="Thao tác trả hàng")
    def return_item_action(self, obj):
        if obj.pk:
            url = reverse('admin:order_item_return_action', args=[obj.pk])
            return format_html(
                '''
                <a class="button" href="#" style="background: #ba2121; color: white; padding: 4px 10px; border-radius: 4px; text-decoration: none;"
                   onclick="var tr = this.closest('tr');
                            var input = tr.querySelector('input[name$=&quot;returned_quantity&quot;]');
                            var qty = input ? input.value : 0;
                            window.location.href = '{}?qty=' + encodeURIComponent(qty);
                            return false;">Trả hàng</a>
                ''',
                url
            )
        return "Lưu trước khi thao tác"


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = [
        'id', 'full_name', 'phone', 'final_price_display',
        'created_at_display', 'order_status', 'points_display',
        'print_status_display', 'order_actions'
    ]
    list_filter = ['created_at', 'is_printed', 'full_name', 'points_awarded']
    search_fields = ['full_name', 'phone', 'id']
    readonly_fields = ['created_at', 'total_price', 'id', 'printed_at']
    inlines = [OrderItemInline]
    list_per_page = 50
    date_hierarchy = 'created_at'

    @admin.display(description="Ngày có đơn hàng")
    def created_at_display(self, obj):
        if not getattr(obj, 'created_at', None):
            return "-"
        try:
            return timezone.localtime(obj.created_at).strftime("%H:%M %d/%m/%Y")
        except Exception:
            return str(obj.created_at)

    def order_status(self, obj):
        now = timezone.now()
        time_diff = now - obj.created_at
        limit = Order.ORDER_DURATION

        if time_diff < limit:
            remaining_days = (limit - time_diff).days + 1
            status, color = f"Còn {remaining_days} ngày", "#ffc107"
        else:
            status, color = "Hoàn thành", "#51cf66"

        return format_html('<span style="color: {}; font-weight: bold;">{}</span>', color, status)

    order_status.short_description = "Đang Xử Lý"

    @admin.display(description="Trạng thái in")
    def print_status_display(self, obj):
        if getattr(obj, 'is_printed', False):
            time_str = obj.printed_at.strftime('%H:%M %d/%m/%Y') if obj.printed_at else ""
            return format_html(
                '<span style="color: #fff; background-color: #2b8a3e; padding: 4px 8px; border-radius: 3px; font-weight: bold;" title="In lúc: {}">✓ Đã in</span>',
                time_str
            )
        return mark_safe(
            '<span style="color: #fff; background-color: #adb5bd; padding: 4px 8px; border-radius: 3px; font-weight: bold;">Chưa in</span>'
        )

    @admin.display(description="Điểm cộng")
    def points_display(self, obj):
        points = obj.awarded_points if (obj.awarded_points and obj.awarded_points > 0) else obj.calculate_points()
        if points > 0:
            if getattr(obj, 'points_awarded', False):
                return format_html(
                    '<span style="color: #2b8a3e; font-weight: bold; background-color: #e8f5e9; padding: 4px 8px; border-radius: 3px;">✓ {} điểm</span>',
                    points
                )
            return format_html(
                '<span style="color: #d97706; font-weight: bold; background-color: #fef3c7; padding: 4px 8px; border-radius: 3px;">x {} điểm</span>',
                points
            )
        return "-"

    def order_actions(self, obj):
        now = timezone.now()
        time_diff = now - obj.created_at

        print_url = reverse('admin:order_print_action', args=[obj.pk])
        print_btn = f'<a class="button" href="{print_url}" target="_blank" style="background: #417690; color: white; margin-right: 4px; padding: 4px 10px; border-radius: 4px; text-decoration: none;">In đơn</a>'

        if time_diff < Order.ORDER_DURATION:
            return_url = reverse('admin:order_return_action', args=[obj.pk])
            return_btn = f'<a class="button" href="{return_url}" style="background: #ba2121; color: white; padding: 4px 10px; border-radius: 4px; text-decoration: none;">Hoàn trả</a>'
            return format_html('{} {}', mark_safe(print_btn), mark_safe(return_btn))
        return format_html('{}', mark_safe(print_btn))

    order_actions.short_description = "Hành động"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<path:object_id>/return-action/', self.admin_site.admin_view(self.return_action_view),
                 name='order_return_action'),
            path('<path:object_id>/print/', self.admin_site.admin_view(self.print_order_view),
                 name='order_print_action'),
            path('order-item/<path:item_id>/return/', self.admin_site.admin_view(self.return_order_item_view),
                 name='order_item_return_action'),
        ]
        return custom_urls + urls

    def print_order_view(self, request, object_id):
        order = get_object_or_404(Order, pk=object_id)
        if not order.is_printed:
            order.is_printed = True
            order.printed_at = timezone.now()
            order.save(update_fields=['is_printed', 'printed_at'])

        config = ShopConfiguration.get_config()
        items = order.items.select_related('product').all()

        # Bảo mật XSS bằng cách escape các chuỗi đầu vào từ dữ liệu người dùng
        safe_title = escape(config.title)
        safe_address = escape(config.address)
        safe_phone = escape(config.phone)
        safe_full_name = escape(order.full_name)
        safe_order_phone = escape(order.phone)
        safe_order_address = escape(order.address)
        safe_note = escape(order.note or 'Không có')

        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>In đơn hàng #{order.id}</title>
            <style>
                body {{ font-family: Arial, sans-serif; font-size: 14px; color: #333; margin: 0; padding: 20px; }}
                .invoice-box {{ max-width: 750px; margin: auto; padding: 25px; border: 1px solid #ddd; background: #fff; }}
                .header {{ text-align: center; margin-bottom: 20px; border-bottom: 2px solid #eee; padding-bottom: 15px; }}
                .header h2 {{ margin: 0; color: #2c3e50; font-size: 22px; }}
                .header p {{ margin: 5px 0; color: #555; font-size: 13px; }}
                .info-section {{ margin-bottom: 20px; font-size: 14px; }}
                .info-section table {{ width: 100%; border-collapse: collapse; }}
                .info-section td {{ padding: 6px; vertical-align: top; }}
                table.items-table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
                table.items-table th, table.items-table td {{ border: 1px solid #ccc; padding: 8px 10px; text-align: left; font-size: 13px; }}
                table.items-table th {{ background-color: #f1f1f1; }}
                .text-right {{ text-align: right; }}
                .totals {{ margin-top: 15px; float: right; width: 320px; }}
                .totals table {{ width: 100%; border-collapse: collapse; }}
                .totals td {{ padding: 6px; font-size: 14px; }}
                .print-btn {{ text-align: center; margin-top: 40px; clear: both; }}
                .print-btn button {{ background: #417690; color: white; border: none; padding: 12px 25px; font-size: 16px; cursor: pointer; border-radius: 4px; font-weight: bold; }}
                .print-btn button:hover {{ background: #205067; }}
                @media print {{
                    .print-btn {{ display: none; }}
                    body {{ padding: 0; }}
                    .invoice-box {{ border: none; padding: 0; }}
                }}
            </style>
        </head>
        <body>
            <div class="invoice-box">
                <div class="header">
                    <h2>{safe_title}</h2>
                    <p>Địa chỉ: {safe_address} | Hotline: {safe_phone}</p>
                    <h3 style="margin-top: 10px; margin-bottom: 5px; color: #333;">PHIẾU GIAO HÀNG / HÓA ĐƠN BÁN HÀNG</h3>
                    <p>Mã đơn hàng: <b>#{order.id}</b> | Ngày đặt: {order.created_at.strftime('%H:%M %d/%m/%Y')}</p>
                </div>

                <div class="info-section">
                    <table>
                        <tr>
                            <td style="width: 50%;">
                                <b>Khách hàng:</b> {safe_full_name}<br>
                                <b>Số điện thoại:</b> {safe_order_phone}<br>
                            </td>
                            <td>
                                <b>Địa chỉ nhận hàng:</b> {safe_order_address}<br>
                                <b>Ghi chú đơn:</b> <span style="color: #d9534f;">{safe_note}</span>
                            </td>
                        </tr>
                    </table>
                </div>

                <table class="items-table">
                    <thead>
                        <tr>
                            <th style="width: 40px; text-align: center;">STT</th>
                            <th style="width: 100px;">Mã SP</th>
                            <th>Tên sản phẩm</th>
                            <th style="width: 70px; text-align: center;">SL</th>
                            <th style="text-align: right;">Đơn giá</th>
                            <th style="text-align: right;">Thành tiền</th>
                        </tr>
                    </thead>
                    <tbody>
        """

        item_rows = ""
        for index, item in enumerate(items, 1):
            base_price = item.price or 0
            discount = item.discount_per_unit or 0
            unit_price = base_price - discount
            subtotal = unit_price * item.quantity
            price_str = f"{unit_price:,.0f}".replace(",", ".") + "đ"
            subtotal_str = f"{subtotal:,.0f}".replace(",", ".") + "đ"

            product_code = escape(item.product.code if item.product and item.product.code else '---')
            product_name = escape(item.product.name if item.product else 'Sản phẩm')

            item_rows += f"""
                                <tr>
                                    <td style="text-align: center;">{index}</td>
                                    <td>{product_code}</td>
                                    <td>{product_name}</td>
                                    <td style="text-align: center;">{item.quantity}</td>
                                    <td style="text-align: right;">{price_str}</td>
                                    <td style="text-align: right;">{subtotal_str}</td>
                                </tr>
                    """

        html_content += item_rows
        final_price_str = f"{order.final_price:,.0f}".replace(",", ".") + "đ"

        html_content += f"""
                    </tbody>
                </table>
                <div style="clear: both;"></div>
                <div class="totals">
                    <table>                        
                        <tr style="border-top: 2px solid #333;">
                            <td><b>Thanh Toán:</b></td>
                            <td class="text-right" style="font-size: 16px; color: #d9534f;"><b>{final_price_str}</b></td>
                        </tr>
                    </table>
                </div>
                <div style="clear: both;"></div>
                <div class="print-btn">
                    <button onclick="window.print();">🖨️ In Hóa Đơn Ngay</button>
                </div>
            </div>
        </body>
        </html>
        """
        return HttpResponse(html_content)

    def return_order_item_view(self, request, item_id):
        order_item = get_object_or_404(OrderItem.objects.select_related('order', 'product'), pk=item_id)
        order = order_item.order

        if timezone.now() - order.created_at >= Order.ORDER_DURATION:
            messages.error(request, f"Không thể trả hàng: Đơn hàng '{order.id}' đã quá hạn.")
            return redirect(reverse('admin:shop_order_change', args=[order.pk]))

        try:
            returned_qty = int(request.GET.get('qty', 0))
        except (ValueError, TypeError):
            returned_qty = 0

        if returned_qty <= 0 or returned_qty > order_item.quantity:
            messages.error(request, "Số lượng trả hàng không hợp lệ!")
            return redirect(reverse('admin:shop_order_change', args=[order.pk]))

        try:
            with transaction.atomic():
                product = order_item.product
                base_price = order_item.price or 0
                discount_per_unit = order_item.discount_per_unit or 0
                final_unit_price = base_price - discount_per_unit

                product.stock += returned_qty
                product.save()

                refund_total_amount = base_price * returned_qty
                refund_final_amount = final_unit_price * returned_qty
                refund_points_deduct = discount_per_unit * returned_qty

                order.total_price = max(Decimal(0), order.total_price - refund_total_amount)
                order.final_price = max(Decimal(0), order.final_price - refund_final_amount)

                if order.applied_points > 0 and refund_points_deduct > 0:
                    points_to_refund = int(refund_points_deduct)
                    order.applied_points = max(0, order.applied_points - points_to_refund)

                    clean_phone = str(order.phone).strip() if order.phone else ""
                    if clean_phone:
                        customer = Customer.objects.filter(phone=clean_phone).first()
                        if customer:
                            customer.points += points_to_refund
                            customer.save(update_fields=['points'])

                amount_for_points = order.final_price if order.final_price > 0 else order.total_price
                order.awarded_points = int(amount_for_points * Decimal('0.01'))
                order.save()

                if returned_qty == order_item.quantity:
                    order_item.delete()
                else:
                    order_item.quantity -= returned_qty
                    order_item.returned_quantity = 0
                    order_item.save()

                messages.success(request, f"Đã xử lý trả thành công {returned_qty} sản phẩm.")
        except Exception as e:
            messages.error(request, f"Lỗi khi xử lý trả hàng: {str(e)}")

        return redirect(reverse('admin:shop_order_change', args=[order.pk]))

    def return_action_view(self, request, object_id):
        order = get_object_or_404(Order, pk=object_id)
        if timezone.now() - order.created_at >= Order.ORDER_DURATION:
            messages.error(request, "Không thể hoàn trả: Đơn hàng đã quá hạn!")
            return redirect('admin:shop_order_changelist')

        try:
            with transaction.atomic():
                order_items = OrderItem.objects.filter(order=order).select_related('product')
                for order_item in order_items:
                    product = order_item.product
                    product.stock += order_item.quantity
                    product.save()

                clean_phone = str(order.phone).strip() if order.phone else ""
                if clean_phone:
                    customer = Customer.objects.filter(phone=clean_phone).first()
                    if customer:
                        if order.applied_points and order.applied_points > 0:
                            customer.points += order.applied_points
                        if getattr(order, 'points_awarded',
                                   False) and order.awarded_points and order.awarded_points > 0:
                            customer.points = max(0, customer.points - order.awarded_points)
                        customer.save(update_fields=['points'])

                order.delete()
                messages.success(request, f"Đã hoàn trả thành công đơn hàng #{object_id}.")
        except Exception as e:
            messages.error(request, f"Lỗi khi hoàn trả đơn hàng: {str(e)}")

        return redirect('admin:shop_order_changelist')

    def get_queryset(self, request):
        return super().get_queryset(request).prefetch_related(
            Prefetch('items', queryset=OrderItem.objects.select_related('product'))
        )

    @admin.display(description="Thanh toán")
    def final_price_display(self, obj):
        if getattr(obj, 'final_price', None) is not None:
            try:
                return f"{obj.final_price:,.0f}".replace(",", ".")
            except Exception:
                return str(obj.final_price)
        return "-"


@admin.register(Customer)
class CustomerAdmin(admin.ModelAdmin):
    list_display = ['full_name', 'phone', 'created_at', 'customer_badge', 'points', 'reset_password_button']
    list_filter = ['created_at', 'phone']
    search_fields = ['full_name', 'phone']
    ordering = ['-created_at']
    list_per_page = 100
    readonly_fields = ['password', 'created_at']
    date_hierarchy = 'created_at'

    def customer_badge(self, obj):
        return format_html(
            '<span style="background-color: #e3f2fd; padding: 3px 10px; border-radius: 3px; font-size: 12px;">ID: {}</span>',
            obj.id
        )

    customer_badge.short_description = "Mã khách"

    def reset_password_button(self, obj):
        url = reverse('admin:customer_reset_password', args=[obj.pk])
        return format_html(
            '<a class="button" style="background: #ba2121; color: white; padding: 4px 8px; border-radius: 4px; text-decoration: none;" href="{}">Cấp lại MK</a>',
            url
        )

    reset_password_button.short_description = "Đổi mật khẩu"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<path:object_id>/reset-password/', self.admin_site.admin_view(self.reset_password_view),
                 name='customer_reset_password'),
        ]
        return custom_urls + urls

    def reset_password_view(self, request, object_id):
        customer = get_object_or_404(Customer, pk=object_id)

        if request.method == 'POST':
            new_password = request.POST.get('new_password')
            if not new_password or len(new_password.strip()) < 4:
                messages.error(request, "Mật khẩu mới quá ngắn hoặc không được để trống!")
            else:
                customer.set_password(new_password.strip())
                customer.save(update_fields=['password'])
                messages.success(request, f"Đã cấp lại mật khẩu thành công cho khách hàng: {customer.full_name} ({customer.phone})")
                return redirect('admin:shop_customer_changelist')

        template_str = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Cấp lại mật khẩu khách hàng</title>
            <meta charset="utf-8">
            <style>
                body { font-family: Arial, sans-serif; background: #f4f6f9; padding: 40px; }
                .box { max-width: 400px; margin: auto; background: #fff; padding: 25px; border: 1px solid #ddd; border-radius: 6px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
                h2 { color: #333; font-size: 18px; margin-top: 0; }
                p { color: #666; font-size: 14px; }
                input[type="text"] { width: 100%; padding: 10px; margin: 15px 0; border: 1px solid #ccc; border-radius: 4px; box-sizing: border-box; }
                button { background: #417690; color: white; border: none; padding: 10px 15px; font-size: 14px; cursor: pointer; border-radius: 4px; font-weight: bold; }
                button:hover { background: #205067; }
                a { margin-left: 10px; color: #666; text-decoration: none; }
            </style>
        </head>
        <body>
            <div class="box">
                <h2>Cấp lại mật khẩu cho: {{ customer.full_name }}</h2>
                <p>Số điện thoại: <b>{{ customer.phone }}</b></p>
                <form method="post">
                    {% csrf_token %}
                    <label>Nhập mật khẩu mới:</label>
                    <input type="text" name="new_password" placeholder="Nhập mật khẩu mới..." required autofocus>
                    <div>
                        <button type="submit">Lưu mật khẩu mới</button>
                        <a href="{% url 'admin:shop_customer_changelist' %}">Quay lại</a>
                    </div>
                </form>
            </div>
        </body>
        </html>
        """
        t = Template(template_str)
        context = RequestContext(request, {'customer': customer})
        return HttpResponse(t.render(context))


class SafeUserAdmin(DjangoUserAdmin):
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)

        class _SafeForm(form):
            def clean(self_inner):
                cleaned = super().clean()
                email = cleaned.get('email', '')
                will_be_superuser = (obj is not None and getattr(obj, 'is_superuser', False)) or cleaned.get('is_superuser', False)

                if will_be_superuser and (email is None or str(email).strip() == ""):
                    raise AdminValidationError("Cannot clear email address of a superuser.")
                return cleaned

        return _SafeForm


try:
    admin.site.unregister(User)
except Exception:
    pass
admin.site.register(User, SafeUserAdmin)


@admin.register(SalesReport)
class SalesReportAdmin(admin.ModelAdmin):
    change_list_template = "admin/shop/reports.html"

    def get_model_perms(self, request):
        return {'add': False, 'change': False, 'delete': False, 'view': True}

    def has_view_permission(self, request, obj=None):
        return request.user.is_active and request.user.is_staff

    def changelist_view(self, request, extra_context=None):
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        group = request.GET.get('group', 'month')

        try:
            dt_from = timezone.make_aware(datetime.strptime(date_from, "%Y-%m-%d")) if date_from else timezone.localtime(timezone.now()).replace(hour=0, minute=0, second=0, microsecond=0)
            dt_to = timezone.make_aware(datetime.combine(datetime.strptime(date_to, "%Y-%m-%d"), datetime.max.time())) if date_to else timezone.localtime(timezone.now())
        except Exception:
            dt_from = timezone.localtime(timezone.now()).replace(hour=0, minute=0, second=0, microsecond=0)
            dt_to = timezone.localtime(timezone.now())

        sold_qty_expr = ExpressionWrapper(F('quantity') - F('returned_quantity'), output_field=DecimalField())
        unit_price_expr = ExpressionWrapper(F('price') - F('discount_per_unit'), output_field=DecimalField(max_digits=18, decimal_places=2))
        cost_per_unit_expr = ExpressionWrapper(Coalesce(F('import_price'), F('product__import_price')), output_field=DecimalField(max_digits=18, decimal_places=2))

        revenue_expr = ExpressionWrapper(unit_price_expr * sold_qty_expr, output_field=DecimalField(max_digits=18, decimal_places=2))
        cogs_expr = ExpressionWrapper(cost_per_unit_expr * sold_qty_expr, output_field=DecimalField(max_digits=18, decimal_places=2))
        profit_expr = ExpressionWrapper((unit_price_expr - cost_per_unit_expr) * sold_qty_expr, output_field=DecimalField(max_digits=18, decimal_places=2))

        items_qs = OrderItem.objects.filter(order__created_at__gte=dt_from, order__created_at__lte=dt_to).select_related('product', 'order')

        agg = items_qs.aggregate(
            total_revenue=Sum(revenue_expr),
            total_cogs=Sum(cogs_expr),
            total_profit=Sum(profit_expr),
        )

        total_revenue = agg.get('total_revenue') or Decimal('0')
        total_cogs = agg.get('total_cogs') or Decimal('0')
        total_profit = agg.get('total_profit') or Decimal('0')
        total_orders = Order.objects.filter(created_at__gte=dt_from, created_at__lte=dt_to).count()

        inv_agg = Product.objects.aggregate(inventory_value=Sum(ExpressionWrapper(F('stock') * F('import_price'), output_field=DecimalField(max_digits=18, decimal_places=2))))
        inventory_value = inv_agg['inventory_value'] or Decimal('0')

        period_field = TruncDay('order__created_at') if group == 'day' else TruncMonth('order__created_at')
        period_qs = items_qs.annotate(period=period_field).values('period').annotate(
            period_revenue=Sum(revenue_expr),
            period_cogs=Sum(cogs_expr),
            period_profit=Sum(profit_expr),
        ).order_by('period')

        labels, rev_data, cogs_data, profit_data = [], [], [], []
        for row in period_qs:
            p = row.get('period')
            if not p:
                continue
            labels.append(p.strftime('%d/%m') if group == 'day' else p.strftime('%m/%Y'))
            rev_data.append(int(row.get('period_revenue') or 0))
            cogs_data.append(int(row.get('period_cogs') or 0))
            profit_data.append(int(row.get('period_profit') or 0))

        top_qs = (
            items_qs
            .values('product__id', 'product__name', 'product__code')
            .annotate(qty_sold=Sum(sold_qty_expr), revenue=Sum(revenue_expr))
            .filter(Q(product__id__isnull=False) & (Q(qty_sold__gt=0) | Q(revenue__gt=0)))
            .order_by('-qty_sold')[:30]
        )

        def fmt_money(v):
            try:
                return f"{int(v):,}".replace(",", ".") + "đ"
            except Exception:
                return "0đ"

        def fmt_int(v):
            try:
                return f"{int(v):,}".replace(",", ".")
            except Exception:
                return "0"

        top_products = []
        for row in top_qs:
            pid = row.get('product__id')
            qty = int(row.get('qty_sold') or 0)
            rev = int(row.get('revenue') or 0)
            top_products.append({
                'id': pid,
                'name': row.get('product__name') or "—",
                'code': row.get('product__code') or "",
                'qty': qty,
                'qty_display': fmt_int(qty),
                'revenue': rev,
                'revenue_display': fmt_money(rev),
                'url': reverse('admin:shop_product_change', args=[pid]) if pid else '#',
            })

        date_from_iso = dt_from.date().isoformat() if hasattr(dt_from, 'date') else ''
        date_to_iso = dt_to.date().isoformat() if hasattr(dt_to, 'date') else ''

        try:
            df_str, dt_str = dt_from.strftime('%d/%m/%Y'), dt_to.strftime('%d/%m/%Y')
            date_range_display = df_str if df_str == dt_str else f"{df_str} — {dt_str}"
        except Exception:
            date_range_display = ''

        export = request.GET.get('export')
        if export in ('xlsx', 'excel'):
            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
                from openpyxl.chart import BarChart, Reference
                from openpyxl.styles import Border, Side
                from openpyxl.chart.label import DataLabelList
            except Exception:
                messages.error(request, "openpyxl không được cài đặt; không thể xuất Excel.")
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Bảng Tóm Tắt"
                ws.append(["Báo Cáo Tài Chính"])
                ws.append([])
                ws.append(["Khoảng thời gian", date_range_display or ""])
                ws.append(["Tổng Doanh thu", float(total_revenue)])
                ws.cell(row=ws.max_row, column=2).number_format = '#,##0'
                ws.append(["Tổng Giá vốn", float(total_cogs)])
                ws.cell(row=ws.max_row, column=2).number_format = '#,##0'
                ws.append(["Tổng Lợi nhuận", float(total_profit)])
                ws.cell(row=ws.max_row, column=2).number_format = '#,##0'
                ws.append(["Số đơn hàng", int(total_orders)])
                ws.cell(row=ws.max_row, column=2).number_format = '#,##0'
                ws.append(["Tổng Giá Tồn Kho", float(inventory_value)])
                ws.cell(row=ws.max_row, column=2).number_format = '#,##0'

                ws2 = wb.create_sheet(title="Biểu Đồ Doanh Thu")
                ws2.append(["Thời Gian", "Doanh Thu", "Giá Vốn", "Lợi Nhuận", "Tỷ Lệ Giá Vốn (%)", "Bin Lợi Nhuận (%)"])

                for i, label in enumerate(labels):
                    rev = rev_data[i] if i < len(rev_data) else 0
                    cogs = cogs_data[i] if i < len(cogs_data) else 0
                    prof = profit_data[i] if i < len(profit_data) else 0
                    cogs_pct = round((cogs / rev) * 100, 2) if rev else 0
                    prof_pct = round((prof / rev) * 100, 2) if rev else 0

                    ws2.append([str(label), rev, cogs, prof, cogs_pct, prof_pct])
                    row_idx = ws2.max_row
                    for col_num in range(2, 5):
                        ws2.cell(row=row_idx, column=col_num).number_format = '#,##0'
                    for col_num in range(5, 7):
                        ws2.cell(row=row_idx, column=col_num).number_format = '0.00'

                try:
                    if len(labels) >= 1:
                        chart = BarChart()
                        chart.type, chart.style = "col", 10
                        chart.title = "Doanh thu - Giá vốn - Lợi nhuận"
                        chart.y_axis.title, chart.x_axis.title = 'VNĐ', 'Thời gian'

                        data_ref = Reference(ws2, min_col=2, min_row=1, max_col=4, max_row=1 + len(labels))
                        chart.add_data(data_ref, titles_from_data=True)
                        chart.set_categories(Reference(ws2, min_col=1, min_row=2, max_row=1 + len(labels)))

                        chart.dataLabels = DataLabelList()
                        chart.dataLabels.showVal = True
                        chart.dataLabels.showSerName = False
                        chart.dataLabels.showCatName = False
                        chart.width, chart.height = 24, 12

                        web_colors = ["3B82F6", "EF4444", "10B981"]
                        for idx, series in enumerate(chart.series):
                            if idx < len(web_colors):
                                series.graphicalProperties.solidFill = web_colors[idx]

                        ws2.add_chart(chart, "H2")
                except Exception:
                    pass

                ws3 = wb.create_sheet(title="Sản Phẩm Bán Chạy")
                ws3.append(["Thứ hạng", "Tên Sản Phẩm", "Số Lượng Bán", "Doanh Thu"])
                for idx, p in enumerate(top_products, start=1):
                    ws3.append([idx, p.get('name') or "", int(p.get('qty') or 0), int(p.get('revenue') or 0)])
                    row_idx = ws3.max_row
                    ws3.cell(row=row_idx, column=1).number_format = '#,##0'
                    ws3.cell(row=row_idx, column=3).number_format = '#,##0'
                    ws3.cell(row=row_idx, column=4).number_format = '#,##0'

                thin_border = Border(left=Side(style='thin', color='888888'), right=Side(style='thin', color='888888'), top=Side(style='thin', color='888888'), bottom=Side(style='thin', color='888888'))
                for wsx in (ws, ws2, ws3):
                    if wsx:
                        for row in wsx.iter_rows(min_row=1, max_row=wsx.max_row, min_col=1, max_col=wsx.max_column):
                            for cell in row:
                                if cell.value is not None:
                                    cell.border = thin_border
                        for col in range(1, (wsx.max_column or 1) + 1):
                            try:
                                wsx.column_dimensions[get_column_letter(col)].width = 18
                            except Exception:
                                pass

                stream = io.BytesIO()
                wb.save(stream)
                stream.seek(0)
                fname = f"financial_report_{group}_{date_from_iso}_{date_to_iso}.xlsx"
                response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                response['Content-Disposition'] = f'attachment; filename="{fname}"'
                return response

        context = {
            'title': 'Báo Cáo Tài Chính',
            'total_revenue_display': fmt_money(total_revenue),
            'total_cogs_display': fmt_money(total_cogs),
            'total_profit_display': fmt_money(total_profit),
            'inventory_value_display': fmt_money(inventory_value),
            'total_orders_display': fmt_int(total_orders),
            'date_from': date_from_iso,
            'date_to': date_to_iso,
            'date_range_display': date_range_display,
            'chart_labels_json': json.dumps(labels),
            'chart_revenue_json': json.dumps(rev_data),
            'chart_cogs_json': json.dumps(cogs_data),
            'chart_profit_json': json.dumps(profit_data),
            'chart_group': group,
            'top_products': top_products,
        }
        return render(request, self.change_list_template, context)
